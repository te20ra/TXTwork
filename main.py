import openpyxl
from openpyxl.utils import get_column_letter
from tkinter import *
from tkinter import ttk
from tkinter import filedialog

FILENAME_XLSX = ''
FILENAME_XML = ''
TABLE = {"number": [],
         "code": [],
         'name': [],
         'count': [],
         'price': [],
         'price_without_nds': [],
         'price_with_nds': [],
         'nds': [],
         'GTD': [],
         'numberUPD': '',
         'dateUPD': ''}


def shortcut_name(name):
    shortname = name[name.rfind('/')+1:]
    if len(shortname) < 30:
        return shortname
    else:
        shortname = shortname[:13] + '...' + shortname[-13:]
        return shortname
def openfile_xlsx(): # функция для открытия даиалогового окна выбора файла
    global FILENAME_XLSX
    FILENAME_XLSX = filedialog.askopenfilename()
    short_filename = shortcut_name(FILENAME_XLSX)
    label_filename = Label(window, text=f'Файл Excel: {short_filename}')
    label_filename.grid(column=1, row=0)


def openfile_xml():
    global FILENAME_XML
    FILENAME_XML = filedialog.askopenfilename()
    short_filename = shortcut_name(FILENAME_XML)
    label_filename = Label(window, text=f'Файл XML: {short_filename}') # создается текст, где прописан путь к
    label_filename.grid(column=1, row=1)#label_filename.grid(column=1, row=0, padx=(50,0),pady=(50,0))

def read_book():
    workbook = openpyxl.load_workbook(FILENAME_XLSX, data_only=True)  # загрузили книгу
    sheets_list = workbook.sheetnames  # выгрузили имена листов
    sheet_active = workbook[sheets_list[2]]  # выбрали 3 лист
    row_max = sheet_active.max_row
    keys = list(TABLE.keys())  # вывод всех ключей в список
    count = 0
    for col in [3, 4, 5, 6, 11, 10, 7, 9, 15]:
        key = keys[count]  # выбор ключа
        for row in range(2, row_max + 1):
            cell_letter = str(get_column_letter(col)) + str(row)  # ячейка
            val = sheet_active[cell_letter].internal_value  # значение ячейки
            if val is None:
                pass
            else:
                TABLE[key].append(val)  # добавили значение ячейки
        count += 1
    TABLE['numberUPD'] = str(sheet_active['A2'].internal_value)
    s = str(sheet_active['B2'].internal_value)
    TABLE['dateUPD'] = s[8:10] + '.' + s[5:7] + '.' + s[0:4]



def table_in_line():
    lengt = int(TABLE['number'][-1])
    line ='<ТаблСчФакт>'
    for i in range(lengt):
        line += f'<СведТов НомСтр="{TABLE["number"][i]}" ' \
                f'НаимТов="{TABLE["name"][i]}"' \
                f' ОКЕИ_Тов="796" КолТов="{TABLE["count"][i]}" ЦенаТов="{TABLE["price"][i]}" СтТовБезНДС="' \
                f'{TABLE["price_without_nds"][i]}" ' \
                f'НалСт="20%" ' \
                f'СтТовУчНал="{TABLE["price_with_nds"][i]}"><Акциз><БезАкциз>без ' \
                f'акциза</БезАкциз></Акциз><СумНал><СумНал>{TABLE["nds"][i]}</СумНал></СумНал>' \
                f'<СвТД КодПроисх="156" НомерТД="{TABLE["GTD"][i]}" />' \
                f'<ДопСведТов ПрТовРаб="1" КодТов="{TABLE["code"][i]}" НаимЕдИзм="шт" КрНаимСтрПр="КИТАЙ" НадлОтп="0" /></СведТов>'
    line += f'<ВсегоОпл СтТовБезНДСВсего="{sum(TABLE["price_without_nds"])}" СтТовУчНалВсего="{sum(TABLE["price_with_nds"])}"><СумНалВсего><СумНал>{sum(TABLE["nds"])}</СумНал></СумНалВсего></ВсегоОпл></ТаблСчФакт>'
    return line

def change_line():
    FILENAME_NEW_XML = FILENAME_XML[:-8] + str(int(FILENAME_XML[-8:-4]) + 1) + '.xml'

    with open(FILENAME_XML, 'r') as main_xml, open(FILENAME_NEW_XML, 'w') as new_xml:
        input_line = main_xml.read()

        num1_start = input_line.find('ВерсФорм')-6
        num1_end = input_line.find('ВерсФорм')-2
        num1 = str(int(input_line[num1_start:num1_end]) + 1)
        output_line = input_line[:num1_start] + num1

        upd1_start = input_line.find('НомерСчФ=')
        codeOKV = input_line[input_line.find('КодОКВ') + 8:input_line.find(' ДатаСчФ') - 1]
        output_line += input_line[num1_end:upd1_start] + 'НомерСчФ="' + TABLE['numberUPD'] + '" КодОКВ="' + codeOKV + \
                       '" ДатаСчФ="' + TABLE['dateUPD'] + '">'

        upd2_start = input_line.find('НомДокОтгр') + 12
        output_line += input_line[input_line.find('<ИспрСчФ ДефНомИспрСчФ'):upd2_start] + TABLE['numberUPD'] + \
            '" ДатаДокОтгр="' + TABLE['dateUPD'] + '" /><ИнфПолФХЖ1></ИнфПолФХЖ1></СвСчФакт>'
        output_line += table_in_line()


        output_line += '<СвПродПер><СвПер СодОпер="Товары переданы"><ОснПер ДатаОсн="' + TABLE['dateUPD'] + \
            '" НаимОсн="Уведомление о выкупе" НомОсн="' + TABLE['numberUPD'] + '" />'
        output_line += input_line[input_line.find('<СвПерВещи'):]

        new_xml.write(output_line)

def start():
    global TABLE
    read_book()
    change_line()
    label_end = Label(window, text=f'Выполнено')
    label_end.grid(column=0, row=3)



window = Tk() # создается окно интрефейса
window.title("Excel to XML")
window.geometry("400x300")

for c in range(10): window.columnconfigure(index=c, weight=10)
for r in range(10): window.rowconfigure(index=r, weight=5)


button_chose_xlsx = Button(window, text='Выбрать ексель файл', command=openfile_xlsx) #кнопка с функцией откртия файла
button_chose_xlsx.grid(column=0, row=0)  #button_chose.grid(column=0, row=0, padx=(50,0), pady=(50,0))

button_chose_xml = Button(window, text='Выбрать xml файл', command=openfile_xml) #кнопка с функцией откртия файла
button_chose_xml.grid(column=0, row=1)  #button_chose.grid(column=0, row=0, padx=(50,0), pady=(50,0))

button_start = Button(window, text='Выполнить', command=start)
button_start.grid(column=0, row=2)  #button_start.grid(column=0, row=1, padx=(10,0), pady=(50,0))




window.mainloop()

#its reary on 25.06.2023 18:45